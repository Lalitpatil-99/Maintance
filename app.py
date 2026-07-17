import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os

EXCEL_FILE = "Society_Maintenance_and_Expenses_Tracker.xlsx"

st.set_page_config(page_title="Arjun Apartments Portal", layout="wide", page_icon="🏢")

# Custom Styling
st.markdown("""
    <style>
    .main-title { font-size: 32px; font-weight: bold; color: #1B365D; text-align: center; margin-bottom: 20px; }
    .kpi-box { background-color: #F0F4F8; padding: 20px; border-radius: 8px; border-left: 5px solid #1B365D; text-align: center; }
    .kpi-val { font-size: 24px; font-weight: bold; color: #1B365D; }
    </style>
""", unsafe_allow_html=True)

st.markdown("<div class='main-title'>🏢 Arjun Apartments Management Portal</div>", unsafe_allow_html=True)

# Helper function to load data safely
def load_data():
    if not os.path.exists(EXCEL_FILE):
        st.error(f"Excel file '{EXCEL_FILE}' not found! Please ensure it is in the same directory.")
        return None, None
    
    try:
        df_main = pd.read_excel(EXCEL_FILE, sheet_name="Maintenance Detail", skiprows=1)
        # Drop total row if exists
        df_main = df_main[df_main['Sr. No.'].notna()]
        df_main = df_main[df_main['Flat/Shop No.'].str.contains('Total|total') == False]
        
        df_exp = pd.read_excel(EXCEL_FILE, sheet_name="Expenses", skiprows=1)
        df_exp = df_exp[df_exp['Sr. No.'].notna()]
        return df_main, df_exp
    except Exception as e:
        st.error(f"Error loading sheets: {e}")
        return None, None

df_main, df_exp = load_data()

if df_main is not None and df_exp is not None:
    # Sidebar Navigation
    menu = st.sidebar.radio("Navigation Menu", ["📊 Dashboard View", "💰 Record Maintenance Payment", "📉 Log Society Expense"])
    
    # ------------------ MENU 1: DASHBOARD ------------------
    if menu == "📊 Dashboard View":
        st.subheader("Operational Insights")
        
        # Calculations
        total_collected = df_main['Maintenance Amount'].sum()
        total_expenses = df_exp['Amount Paid'].sum()
        net_balance = total_collected - total_expenses
        
        total_units = len(df_main)
        paid_units = len(df_main[df_main['Status'].astype(str).str.strip().str.lower() == 'paid'])

        collection_pct = (paid_units / total_units) * 100 if total_units > 0 else 0
        
        # KPI Cards
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"<div class='kpi-box'>Total Collected<br><span class='kpi-val'>₹{total_collected:,.2f}</span></div>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<div class='kpi-box'>Total Expenses<br><span class='kpi-val'>₹{total_expenses:,.2f}</span></div>", unsafe_allow_html=True)
        with col3:
            st.markdown(f"<div class='kpi-box'>Net Balance<br><span class='kpi-val'>₹{net_balance:,.2f}</span></div>", unsafe_allow_html=True)
        with col4:
            st.markdown(f"<div class='kpi-box'>Collection Rate<br><span class='kpi-val'>{collection_pct:.1f}%</span></div>", unsafe_allow_html=True)
            
        st.write("---")
        
        # Visual Split Tables
        c_left, c_right = st.columns(2)
        with c_left:
            st.markdown("### ⚠️ Current Outstanding / Defaulters")
                        # Convert status column to lowercase string and strip spaces to ensure a perfect match
            df_main['Status_Clean'] = df_main['Status'].astype(str).str.strip().str.lower()
            
            # Match 'pending' regardless of how it is written in Excel (e.g. "Pending", "pending ", "PENDING")
            pending_df = df_main[df_main['Status_Clean'] == 'pending'][['Flat/Shop No.', 'Owner Name', "Receiver's Name"]]
            
            st.dataframe(pending_df.rename(columns={"Receiver's Name": 'Assigned Collector'}), use_container_width=True, hide_index=True)

            
        with c_right:
            st.markdown("### 📈 Expense Overview")
            st.dataframe(df_exp[['Month', 'Description', 'Amount Paid']], use_container_width=True, hide_index=True)

    # ------------------ MENU 2: MAINTENANCE RECORDING ------------------
    elif menu == "💰 Record Maintenance Payment":
        st.subheader("Update Maintenance Payment")
        
        # Select Unit from existing units
        unit_list = df_main['Flat/Shop No.'].dropna().unique().tolist()
        selected_unit = st.selectbox("Select Flat / Shop Number", unit_list)
        
        # Fetch current details if available
        current_row = df_main[df_main['Flat/Shop No.'] == selected_unit].iloc[0]
        
        with st.form("maintenance_form", clear_on_submit=True):
            owner_name = st.text_input("Owner Name", value=str(current_row['Owner Name']) if pd.notna(current_row['Owner Name']) else "")
            mobile = st.text_input("Mobile Number", value=str(current_row['Mobile Number']) if pd.notna(current_row['Mobile Number']) else "")
            
            # Standard pricing check
            default_amt = 500.0 if "Shop" in str(selected_unit) else 700.0
            amount = st.number_input("Amount Paid (₹)", min_value=0.0, value=default_amt)
            
            pay_mode = st.selectbox("Payment Mode", ["Online", "Cash"])
            pay_date = st.date_input("Date of Payment", datetime.date.today())
            receiver = st.text_input("Receiver Name", value="Vivek")
            
            submit_btn = st.form_submit_button("Save Payment Record & Update Excel")
            
            if submit_btn:
                # Openpyxl updates directly into the file formatting setup
                wb = openpyxl.load_workbook(EXCEL_FILE)
                ws = wb["Maintenance Detail"]
                
                # Locate row matching Flat/Shop No.
                target_row = None
                for row in range(3, ws.max_row + 1):
                    if ws.cell(row=row, column=4).value == selected_unit:
                        target_row = row
                        break
                
                if target_row:
                    ws.cell(row=target_row, column=2, value=owner_name)
                    ws.cell(row=target_row, column=3, value=mobile)
                    ws.cell(row=target_row, column=5, value=amount)
                    ws.cell(row=target_row, column=6, value=pay_mode)
                    ws.cell(row=target_row, column=7, value=pay_date.strftime("%d %B %Y"))
                    ws.cell(row=target_row, column=8, value=receiver)
                    
                    wb.save(EXCEL_FILE)
                    st.success(f"🎉 Successfully updated payment for {selected_unit}! The Excel file has been saved.")
                    st.experimental_rerun()
                else:
                    st.error("Unit not found in spreadsheet layout.")

    # ------------------ MENU 3: LOG EXPENSES ------------------
    elif menu == "📉 Log Society Expense":
        st.subheader("Log a New Expense Entry")
        
    
        with st.form("expense_form", clear_on_submit=True):
            exp_month = st.selectbox("Expense Month", ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"], index=6)
            exp_desc = st.text_input("Expense Description (e.g. Lift Maintenance, Water Tanker)")
            exp_amt = st.number_input("Amount Paid (₹)", min_value=1.0, step=100.0)
            exp_date = st.date_input("Date of Payment", datetime.date.today())
            
            submit_exp = st.form_submit_button("Add Expense")
            
            if submit_exp and exp_desc:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                ws = wb["Expenses"]
                
                # Find the Total Row to insert right above it cleanly
                insert_row = ws.max_row
                for row in range(3, ws.max_row + 1):
                    val = str(ws.cell(row=row, column=3).value)
                    if "Total" in val:
                        insert_row = row
                        break
                
                # Insert a clean new row right above total
                ws.insert_rows(insert_row)
                
                # Get dynamic sr number
                prev_sr = ws.cell(row=insert_row-1, column=1).value
                new_sr = int(prev_sr) + 1 if prev_sr and str(prev_sr).isdigit() else 1
                
                # Populating values
                ws.cell(row=insert_row, column=1, value=new_sr)
                ws.cell(row=insert_row, column=2, value=exp_month)
                ws.cell(row=insert_row, column=3, value=exp_desc)
                ws.cell(row=insert_row, column=4, value=exp_amt)
                ws.cell(row=insert_row, column=4).number_format = '₹#,##0.00'
                ws.cell(row=insert_row, column=5, value=exp_date.strftime("%d %B %Y"))
                
                # Update total formula string dynamically
                ws.cell(row=insert_row+1, column=4, value=f"=SUM(D3:D{insert_row})")
                
                wb.save(EXCEL_FILE)
                st.success(f"📉 Added expense: '{exp_desc}' of ₹{exp_amt:,.2f} registered successfully!")
                st.experimental_rerun()
